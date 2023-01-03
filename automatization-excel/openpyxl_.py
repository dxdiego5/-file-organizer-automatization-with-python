from openpyxl import load_workbook
import os
# pegando caminho raiz 
base_path = os.path.abspath(os.path.dirname(__file__))
# caminho completo
path = os.path.join(base_path, "exemplo.xlsx")

wb = load_workbook(path)

# abas = wb.get_sheet_names()
sheet = wb['Sheet1']
sheet.cell(row=2, column=2).value

for i in range(1, sheet.max_row):
    print(sheet.cell(row=i, column=2).value)

sheet.cell(row=2, column=3).value = 100
wb.save("exemplo.xlsx")


#### agrupamento de celulas ####
# sheet.merge_cells('A1:D1')
# sheet.unmerge_cells('A1:D1')

#### add imagem ####
# from openpyxl.drawing.image import Image
# img = Image('catlogo.png')
# wb.add_image(img, 'A1')
# wb.save('exemplo2.xlsx')